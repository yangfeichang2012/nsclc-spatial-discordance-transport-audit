#!/usr/bin/env python3
"""Reconstruct the locked PRJ-2026-015 results from public source files.

Only aggregate results and non-identifying bootstrap draws are written. Yale
coefficients are reconstructed to verify the sealed values; no target model is
refitted or recalibrated.
"""

from __future__ import annotations

import argparse
import csv
import gzip
import json
import math
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import numpy as np
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
GENES = ("CXCL14", "PECAM1", "APOE", "CD68")
SEED = 20260731
N_BOOT = 2000
EXPECTED = {
    "B0_M": 1.2351555460162515,
    "B1_M": 1.2392799449001186,
    "B1_D": -0.023505738171393748,
}


def clean(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if len(text) >= 2 and text[0] == text[-1] == '"':
        text = text[1:-1]
    return text.replace('""', '"').strip()


def number(value: object) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return None
    return parsed if math.isfinite(parsed) else None


def binary(value: object) -> float | None:
    if isinstance(value, bool):
        return float(value)
    parsed = number(value)
    if parsed in (0.0, 1.0):
        return parsed
    if isinstance(value, str):
        token = value.strip().lower()
        if token in {"high", "yes", "true", "positive", "pos"}:
            return 1.0
        if token in {"low", "no", "false", "negative", "neg"}:
            return 0.0
    return None


def robust(raw: np.ndarray) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    medians = np.median(raw, axis=0)
    q25 = np.quantile(raw, 0.25, axis=0, method="linear")
    q75 = np.quantile(raw, 0.75, axis=0, method="linear")
    scales = (q75 - q25) / 1.349
    if np.any(~np.isfinite(scales)) or np.any(scales <= 0):
        raise RuntimeError("invalid cohort-local robust scale")
    z = (raw - medians) / scales
    resistance = 0.5 * (z[:, 0] + z[:, 1])
    response = 0.5 * (z[:, 2] + z[:, 3])
    return medians, scales, resistance - response


def harrell_c(time: np.ndarray, event: np.ndarray, risk: np.ndarray) -> tuple[float, int]:
    concordance = 0.0
    comparable = 0
    for left in range(len(time) - 1):
        for right in range(left + 1, len(time)):
            if time[left] == time[right]:
                continue
            if time[left] < time[right] and event[left] == 1:
                earlier, later = left, right
            elif time[right] < time[left] and event[right] == 1:
                earlier, later = right, left
            else:
                continue
            comparable += 1
            if risk[earlier] > risk[later]:
                concordance += 1.0
            elif risk[earlier] == risk[later]:
                concordance += 0.5
    if comparable == 0:
        raise RuntimeError("no comparable pairs")
    return concordance / comparable, comparable


def cox_score_info(
    beta_value: np.ndarray, design: np.ndarray, time: np.ndarray, event: np.ndarray
) -> tuple[float, np.ndarray, np.ndarray]:
    eta = np.clip(design @ beta_value, -40.0, 40.0)
    weights = np.exp(eta)
    loglik = 0.0
    score = np.zeros(design.shape[1])
    info = np.zeros((design.shape[1], design.shape[1]))
    for event_time in np.unique(time[event == 1]):
        deaths = np.flatnonzero((time == event_time) & (event == 1))
        risk = np.flatnonzero(time >= event_time)
        count = len(deaths)
        s0 = weights[risk].sum()
        s1 = (weights[risk, None] * design[risk]).sum(axis=0)
        s2 = np.einsum("i,ij,ik->jk", weights[risk], design[risk], design[risk])
        e0 = weights[deaths].sum()
        e1 = (weights[deaths, None] * design[deaths]).sum(axis=0)
        e2 = np.einsum("i,ij,ik->jk", weights[deaths], design[deaths], design[deaths])
        loglik += eta[deaths].sum()
        score += design[deaths].sum(axis=0)
        for tied_index in range(count):
            fraction = tied_index / count
            denominator = s0 - fraction * e0
            first = s1 - fraction * e1
            second = s2 - fraction * e2
            mean = first / denominator
            loglik -= math.log(denominator)
            score -= mean
            info += second / denominator - np.outer(mean, mean)
    return loglik, score, info


def fit_cox(design: np.ndarray, time: np.ndarray, event: np.ndarray) -> np.ndarray:
    beta_value = np.zeros(design.shape[1])
    for _ in range(100):
        loglik, score, info = cox_score_info(beta_value, design, time, event)
        step = np.linalg.solve(info, score)
        scale = 1.0
        while scale >= 2**-20:
            candidate = beta_value + scale * step
            candidate_loglik, _, _ = cox_score_info(candidate, design, time, event)
            if np.max(np.abs(candidate)) <= 10 and candidate_loglik >= loglik - 1e-10:
                beta_value = candidate
                break
            scale *= 0.5
        else:
            raise RuntimeError("Cox line search failed")
        _, new_score, _ = cox_score_info(beta_value, design, time, event)
        if np.max(np.abs(scale * step)) < 1e-8 and np.max(np.abs(new_score)) < 1e-6:
            return beta_value
    raise RuntimeError("Cox model failed to converge")


def read_spatial_sheet(
    workbook, sheet: str, id_column: str, genes: tuple[str, str]
) -> dict[str, dict[str, float]]:
    rows = workbook[sheet].iter_rows(values_only=True)
    header = [clean(value) for value in next(rows)]
    lower_header = [value.lower() for value in header]
    requested = [
        id_column,
        "binary_score",
        "PFS_2Years_months",
        "PFS_2Years_index",
        *genes,
    ]
    missing = [column for column in requested if column.lower() not in lower_header]
    if missing:
        raise RuntimeError(f"{sheet} missing {missing}")
    index = {column: lower_header.index(column.lower()) for column in requested}
    result: dict[str, dict[str, float]] = {}
    for row in rows:
        patient = clean(row[index[id_column]])
        if not patient:
            continue
        values = {
            "binary": binary(row[index["binary_score"]]),
            "time": number(row[index["PFS_2Years_months"]]),
            "event": binary(row[index["PFS_2Years_index"]]),
            **{gene: number(row[index[gene]]) for gene in genes},
        }
        if values["binary"] is None or any(values[gene] is None for gene in genes):
            raise RuntimeError(f"{sheet} has an incomplete predictor row")
        if patient in result:
            raise RuntimeError(f"{sheet} duplicate identifier")
        result[patient] = {
            key: (None if value is None else float(value)) for key, value in values.items()
        }
    return result


def load_yale_or_greek(figure6: Path, cohort: str) -> dict[str, Any]:
    if cohort == "Yale":
        tumor_sheet, stroma_sheet, id_column = (
            "source_data_Figure_6b",
            "source_data_Figure_6f",
            "Spot_ID",
        )
        expected_n, expected_events = 34, 20
    elif cohort == "Greek":
        tumor_sheet, stroma_sheet, id_column = (
            "source_data_Figure_6d",
            "source_data_Figure_6h",
            "ROILabel",
        )
        expected_n, expected_events = 61, 45
    else:
        raise ValueError(cohort)
    workbook = load_workbook(figure6, read_only=True, data_only=True)
    try:
        tumor = read_spatial_sheet(workbook, tumor_sheet, id_column, ("CXCL14", "PECAM1"))
        stroma = read_spatial_sheet(workbook, stroma_sheet, id_column, ("APOE", "CD68"))
    finally:
        workbook.close()
    rows = []
    for patient in sorted(set(tumor) & set(stroma)):
        trow, srow = tumor[patient], stroma[patient]
        if (
            trow["event"] is not None
            and srow["event"] is not None
            and trow["event"] != srow["event"]
        ):
            raise RuntimeError(f"{cohort} cross-compartment outcome disagreement")
        if (
            trow["time"] is not None
            and srow["time"] is not None
            and abs(trow["time"] - srow["time"]) > 1e-12
        ):
            raise RuntimeError(f"{cohort} cross-compartment outcome disagreement")
        rows.append(
            (
                [trow["CXCL14"], trow["PECAM1"], srow["APOE"], srow["CD68"]],
                trow["binary"] - srow["binary"],
                trow["time"] if trow["time"] is not None else srow["time"],
                trow["event"] if trow["event"] is not None else srow["event"],
            )
        )
    raw_all = np.asarray([row[0] for row in rows], dtype=float)
    medians, scales, discordance = robust(raw_all)
    complete = [row for row in rows if row[2] is not None and row[3] is not None and row[2] > 0]
    raw = np.asarray([row[0] for row in complete], dtype=float)
    z = (raw - medians) / scales
    d_value = 0.5 * (z[:, 0] + z[:, 1]) - 0.5 * (z[:, 2] + z[:, 3])
    result = {
        "m": np.asarray([row[1] for row in complete]),
        "d": d_value,
        "time": np.asarray([row[2] for row in complete]),
        "event": np.asarray([int(row[3]) for row in complete], dtype=int),
        "n": len(complete),
        "events": sum(int(row[3]) for row in complete),
        "predictor_n": len(rows),
        "medians": medians,
        "scales": scales,
    }
    if result["n"] != expected_n or result["events"] != expected_events:
        raise RuntimeError(f"{cohort} denominator gate failed")
    return result


def parse_geo(path: Path) -> dict[str, dict[str, str | None]]:
    selected: list[list[str]] = []
    with gzip.open(path, "rt", encoding="utf-8", errors="replace", newline="") as handle:
        for row in csv.reader(handle, delimiter="\t"):
            if row and row[0].startswith("!Sample_"):
                selected.append(row)
    indexed: dict[str, list[list[str]]] = defaultdict(list)
    for row in selected:
        indexed[row[0]].append(row)
    titles = indexed["!Sample_title"][0][1:]
    result = {}
    for position, title_value in enumerate(titles):
        characteristics: dict[str, list[str]] = defaultdict(list)
        for row in indexed["!Sample_characteristics_ch1"]:
            value = clean(row[position + 1]) if position + 1 < len(row) else ""
            if value and ":" in value:
                key, item = value.split(":", 1)
                characteristics[key.strip().lower()].append(item.strip())
        patients = sorted(set(characteristics.get("patient id", [])))
        title = clean(title_value)
        result[title] = {
            "patient": patients[0] if len(patients) == 1 else None,
            "compartment": (
                "tumor" if "PanCK pos" in title else "stroma" if "PanCK neg" in title else "other"
            ),
        }
    return result


def load_uq(
    figure6: Path, supplementary: Path, geo_path: Path, normalized_path: Path
) -> dict[str, Any]:
    geo = parse_geo(geo_path)
    workbook = load_workbook(normalized_path, read_only=True, data_only=True)
    expression: dict[str, dict[str, dict[str, list[float]]]] = defaultdict(
        lambda: defaultdict(lambda: defaultdict(list))
    )
    try:
        rows = workbook.active.iter_rows(values_only=True)
        header = [clean(value) for value in next(rows)]
        positions = {gene: header.index(gene) for gene in GENES}
        normalized_rows = 0
        for row in rows:
            title = clean(row[0])
            if not title:
                continue
            normalized_rows += 1
            metadata = geo.get(title)
            if not metadata or metadata["patient"] is None or metadata["compartment"] not in {"tumor", "stroma"}:
                raise RuntimeError("UQ normalized row lacks exact public mapping")
            for gene, position in positions.items():
                value = number(row[position])
                if value is None:
                    raise RuntimeError("UQ anchor missing")
                expression[str(metadata["patient"])][str(metadata["compartment"])][gene].append(value)
    finally:
        workbook.close()
    paired = sorted(
        patient
        for patient, compartments in expression.items()
        if compartments["tumor"] and compartments["stroma"]
    )
    if normalized_rows != 68 or len(paired) != 21:
        raise RuntimeError("UQ expression gate failed")
    anchors = {
        patient: {
            "CXCL14": float(np.median(expression[patient]["tumor"]["CXCL14"])),
            "PECAM1": float(np.median(expression[patient]["tumor"]["PECAM1"])),
            "APOE": float(np.median(expression[patient]["stroma"]["APOE"])),
            "CD68": float(np.median(expression[patient]["stroma"]["CD68"])),
        }
        for patient in paired
    }

    binary_values: dict[str, dict[str, list[float]]] = defaultdict(lambda: defaultdict(list))
    workbook = load_workbook(figure6, read_only=True, data_only=True)
    try:
        for compartment, sheet in (
            ("tumor", "source_data_Figure_6c"),
            ("stroma", "source_data_Figure_6g"),
        ):
            rows = workbook[sheet].iter_rows(values_only=True)
            header = [clean(value) for value in next(rows)]
            binary_index = header.index("binary_score")
            for row in rows:
                title = clean(row[0])
                metadata = geo.get(title)
                if not metadata or metadata["patient"] is None or metadata["compartment"] != compartment:
                    raise RuntimeError("UQ binary row lacks exact public mapping")
                value = binary(row[binary_index])
                if value is None:
                    raise RuntimeError("UQ binary missing")
                binary_values[str(metadata["patient"])][compartment].append(value)
    finally:
        workbook.close()
    binaries = {
        patient: {
            "tumor": float(np.median(compartments["tumor"])),
            "stroma": float(np.median(compartments["stroma"])),
        }
        for patient, compartments in binary_values.items()
        if compartments["tumor"] and compartments["stroma"]
    }

    clinical: dict[str, dict[str, object]] = {}
    workbook = load_workbook(supplementary, read_only=True, data_only=True)
    try:
        rows = workbook["Supplementary Data 2"].iter_rows(values_only=True)
        header = [clean(value) for value in next(rows)]
        columns = {
            name: header.index(name)
            for name in (
                "SAMPLE_ID",
                "Date of 1st Immunotherapy infusion",
                "PFS censored",
                "PFS days",
            )
        }
        for row in rows:
            patient = clean(row[columns["SAMPLE_ID"]])
            if patient:
                clinical[patient] = {
                    "start": row[columns["Date of 1st Immunotherapy infusion"]],
                    "event": binary(row[columns["PFS censored"]]),
                    "days": number(row[columns["PFS days"]]),
                }
    finally:
        workbook.close()

    all_raw = np.asarray([[anchors[p][gene] for gene in GENES] for p in sorted(anchors)])
    medians, scales, _ = robust(all_raw)
    complete = sorted(
        patient
        for patient in set(anchors) & set(binaries) & set(clinical)
        if clinical[patient]["start"] not in (None, "")
        and clinical[patient]["days"] is not None
        and clinical[patient]["event"] is not None
        and float(clinical[patient]["days"]) > 0
    )
    if len(complete) != 19:
        raise RuntimeError("UQ complete-case gate failed")
    raw = np.asarray([[anchors[p][gene] for gene in GENES] for p in complete])
    z = (raw - medians) / scales
    result = {
        "m": np.asarray([binaries[p]["tumor"] - binaries[p]["stroma"] for p in complete]),
        "d": 0.5 * (z[:, 0] + z[:, 1]) - 0.5 * (z[:, 2] + z[:, 3]),
        "time": np.asarray([float(clinical[p]["days"]) / 30.4375 for p in complete]),
        "event": np.asarray([int(clinical[p]["event"]) for p in complete], dtype=int),
        "n": 19,
        "events": 18,
        "predictor_n": 21,
        "medians": medians,
        "scales": scales,
    }
    if int(result["event"].sum()) != 18:
        raise RuntimeError("UQ event gate failed")
    return result


def evaluate(data: dict[str, Any]) -> dict[str, float | int]:
    risk0 = EXPECTED["B0_M"] * data["m"]
    risk1 = EXPECTED["B1_M"] * data["m"] + EXPECTED["B1_D"] * data["d"]
    c0, pairs0 = harrell_c(data["time"], data["event"], risk0)
    c1, pairs1 = harrell_c(data["time"], data["event"], risk1)
    if pairs0 != pairs1:
        raise RuntimeError("comparable-pair mismatch")
    return {"C_B0": c0, "C_B1": c1, "delta_C": c1 - c0, "comparable_pairs": pairs0}


def bootstrap(data: dict[str, Any]) -> list[float]:
    rng = np.random.default_rng(SEED)
    values = []
    for _ in range(N_BOOT):
        index = rng.integers(0, data["n"], data["n"])
        c0, _ = harrell_c(
            data["time"][index], data["event"][index], EXPECTED["B0_M"] * data["m"][index]
        )
        c1, _ = harrell_c(
            data["time"][index],
            data["event"][index],
            EXPECTED["B1_M"] * data["m"][index] + EXPECTED["B1_D"] * data["d"][index],
        )
        values.append(c1 - c0)
    return values


def ci(values: list[float]) -> list[float]:
    return [
        float(np.quantile(values, 0.025, method="linear")),
        float(np.quantile(values, 0.975, method="linear")),
    ]


def write_bootstrap(path: Path, values: list[float]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["replicate", "delta_C"])
        writer.writerows(enumerate(values, start=1))


def write_candidate_universe(supplementary: Path, output: Path) -> dict[str, Any]:
    workbook = load_workbook(supplementary, read_only=True, data_only=True)
    try:
        rows = workbook["Supplementary Data 4"].iter_rows(values_only=True)
        header = [clean(value) for value in next(rows)]
        response_col = header.index("CD4_M1_M2_genes")
        resistance_col = header.index("ProliferatingTumor_Granulocytes_Vessels_genes")
        response: list[str] = []
        resistance: list[str] = []
        for row in rows:
            response_gene = clean(row[response_col])
            resistance_gene = clean(row[resistance_col])
            if response_gene:
                response.append(response_gene)
            if resistance_gene:
                resistance.append(resistance_gene)
    finally:
        workbook.close()
    with output.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["parent_list", "row_position", "gene", "selected_anchor"])
        for list_name, genes in (
            ("response_CD4_M1_M2", response),
            ("resistance_ProliferatingTumor_Granulocytes_Vessels", resistance),
        ):
            for position, gene in enumerate(genes, start=1):
                writer.writerow([list_name, position, gene, gene in GENES])
    return {
        "response_unique_n": len(set(response)),
        "resistance_unique_n": len(set(resistance)),
        "anchor_row_positions_not_ranks": {
            "APOE": response.index("APOE") + 1,
            "CD68": response.index("CD68") + 1,
            "CXCL14": resistance.index("CXCL14") + 1,
            "PECAM1": resistance.index("PECAM1") + 1,
        },
    }


def m_summary(data: dict[str, Any]) -> dict[str, int]:
    def label(value: float) -> str:
        return f"{value:g}"

    return dict(sorted(Counter(label(float(value)) for value in data["m"]).items()))


def outcome_summary(data: dict[str, Any]) -> dict[str, Any]:
    time = data["time"]
    event = data["event"]
    return {
        "definition": "T*=min(observed PFS,24 months); event=1 only for progression/death by 24 months",
        "minimum_months": float(np.min(time)),
        "maximum_months": float(np.max(time)),
        "n_at_exactly_24_months": int(np.sum(time == 24)),
        "events_at_exactly_24_months": int(np.sum(event[time == 24])),
        "maximum_event_time_months": float(np.max(time[event == 1])),
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--inputs", type=Path, default=ROOT / "inputs")
    parser.add_argument("--output", type=Path, default=ROOT / "results")
    args = parser.parse_args()
    args.output.mkdir(parents=True, exist_ok=True)
    figure6 = args.inputs / "41588_2025_2351_MOESM10_ESM.xlsx"
    supplementary = args.inputs / "41588_2025_2351_MOESM5_ESM.xlsx"
    geo = args.inputs / "GSE221733_series_matrix.txt.gz"
    normalized = args.inputs / "GSE221733_4301_CTA_norm.xlsx"
    for path in (figure6, supplementary, geo, normalized):
        if not path.is_file():
            raise FileNotFoundError(f"missing input: {path}; run download_inputs.py first")

    yale = load_yale_or_greek(figure6, "Yale")
    greek = load_yale_or_greek(figure6, "Greek")
    uq = load_uq(figure6, supplementary, geo, normalized)
    beta0 = fit_cox(yale["m"][:, None], yale["time"], yale["event"])
    beta1 = fit_cox(np.column_stack([yale["m"], yale["d"]]), yale["time"], yale["event"])
    reconstructed = {"B0_M": float(beta0[0]), "B1_M": float(beta1[0]), "B1_D": float(beta1[1])}
    for name, expected in EXPECTED.items():
        if not math.isclose(reconstructed[name], expected, rel_tol=0, abs_tol=1e-12):
            raise RuntimeError(f"sealed coefficient mismatch for {name}")

    greek_result = evaluate(greek)
    uq_result = evaluate(uq)
    greek_bootstrap = bootstrap(greek)
    uq_bootstrap = bootstrap(uq)
    greek_result["bootstrap_95_percentile_CI"] = ci(greek_bootstrap)
    uq_result["bootstrap_95_percentile_interval"] = ci(uq_bootstrap)
    expected_targets = {
        "Greek": (
            greek_result,
            {
                "C_B0": 0.6136224801466097,
                "C_B1": 0.5919364691508857,
                "delta_C": -0.02168601099572398,
            },
            greek_result["bootstrap_95_percentile_CI"],
            [-0.0640426761828849, 0.020918298917445496],
        ),
        "UQ": (
            uq_result,
            {
                "C_B0": 0.6697530864197531,
                "C_B1": 0.6296296296296297,
                "delta_C": -0.04012345679012341,
            },
            uq_result["bootstrap_95_percentile_interval"],
            [-0.1558493258334023, 0.08824980650154793],
        ),
    }
    for cohort, (observed, expected_values, observed_interval, expected_interval) in expected_targets.items():
        for name, expected_value in expected_values.items():
            if not math.isclose(float(observed[name]), expected_value, rel_tol=0, abs_tol=1e-12):
                raise RuntimeError(f"{cohort} locked result mismatch for {name}")
        if any(
            not math.isclose(float(left), right, rel_tol=0, abs_tol=1e-12)
            for left, right in zip(observed_interval, expected_interval)
        ):
            raise RuntimeError(f"{cohort} locked interval mismatch")
    write_bootstrap(args.output / "greek_bootstrap_delta_c.csv", greek_bootstrap)
    write_bootstrap(args.output / "uq_descriptive_bootstrap_delta_c.csv", uq_bootstrap)
    candidate = write_candidate_universe(
        supplementary, args.output / "candidate_universe.csv"
    )
    payload = {
        "archive_version": "1.0.1",
        "study_identity": "analysis-plan-locked post-deviation secondary transport audit",
        "outcome": "two-year truncated PFS",
        "preprocessing": "target-cohort median/IQR normalization; coefficients fixed, full pipeline not fixed",
        "uncertainty_conditioning": "conditional on fitted Yale coefficients and observed target-cohort preprocessing",
        "sealed_yale_coefficients": reconstructed,
        "Greek_primary_under_amended_G11": {
            "n": greek["n"],
            "events": greek["events"],
            **greek_result,
        },
        "UQ_descriptive_stress_test": {
            "n": uq["n"],
            "events": uq["events"],
            **uq_result,
        },
        "M_value_counts": {
            "Yale": m_summary(yale),
            "Greek": m_summary(greek),
            "UQ": m_summary(uq),
        },
        "two_year_outcome_audit": {
            "Yale": outcome_summary(yale),
            "Greek": outcome_summary(greek),
            "UQ": outcome_summary(uq),
        },
        "candidate_universe": candidate,
        "bootstrap": {
            "repetitions": N_BOOT,
            "seed_reset_separately_for_each_target": SEED,
            "resampling_unit": "patient",
        },
    }
    (args.output / "reproduced_results.json").write_text(
        json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )
    print(json.dumps(payload, indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
